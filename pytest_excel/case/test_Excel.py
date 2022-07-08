#!/usr/bin/env python
# -*- coding: utf-8 -*-


import allure
import openpyxl
import pytest

from pytest_excel.data_driver.excel_read import read_Excel
from pytest_excel.kw.api_key import Apikey


# 在当前文件中所有用例执行之前执行
def setup_module():
    # 1、定义全局变量
    global excel,sheet,excel_path,ak,all_val
    # 2、初始化Excel文件
    excel_path = './data/api_cases.xlsx'
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel['Sheet1']
    # 3、实例化工具类
    ak = Apikey()
    # 4、参数化变量存储字典
    all_val = {}

# pytest参数化
@pytest.mark.parametrize('value',read_Excel())
def test_01(value):
    # 如果存在自定义标题
    if value[10] is not None:
        # 动态生成标题
        allure.dynamic.title(value[10])

    if value[15] is not None:
        # 动态获取story模块名
        allure.dynamic.story(value[15])

    if value[16] is not None:
        # 动态获取feature模块名
        allure.dynamic.feature(value[16])
    if value[17] is not None:
        # 动态获取备注信息
        allure.dynamic.title(value[10])

    if value[18] is not None:
        # 动态获取级别信息（blocker,critical,normal,minor,trivial）
        allure.dynamic.severity(value[18])

    # print(value)
    r = value[0] + 1
    # print("行数： " + str(r))


    if type(value[0]) is int:
        # 准备需要的测试数据
        # 请求参数
        data = value[5]
        # print(data)
        # 校验字段
        assert_value = value[7]
        # print(assert_value)
        # 预期结果字段
        expect_value = value[8]
        # print(expect_value)
        # 如果存在请求头
        if value[4]:
            # print("存在请求头")
            # 存在请求参数
            if value[5]:
                # print("存在请求参数")
                dict_data = {
                    'url': value[1] + value[2],
                    # eval 将字符串str当做有效的表达式来求值并返回计算结果
                    # 这里直接给headers一个字典值
                    'headers':eval(value[4]),
                    value[6]: eval(data)
                }
                print(dict_data)
            # 不存在请求参数
            else:
                # print('不存在请求参数')
                dict_data = {
                    'url': value[1] + value[2],
                    'headers': eval(value[4])
                }
                print(dict_data)

        # 不存在请求头
        else:
            # print("不存在请求头")
            # 存在请求参数
            if value[5]:
                # print('存在请求参数')
                dict_data = {
                    'url': value[1] + value[2],
                    value[6]: eval(data)
                }
                print(dict_data)
            # 不存在请求参数
            else:
                # print('不存在请求参数')
                dict_data = {
                    'url': value[1] + value[2]
                }
                print(dict_data)

        """
            常规的参数传递：
            requests.post(url="",data="",headers="")
            封装时，参数可以直接用**kwargs
            requests.post(**kwargs)
        """
        # 使用反射模拟请求
        res = getattr(ak,value[3])(**dict_data)

        # 打印响应报文
        print(res.text)

        # ===============JSON提取，多参数版==========
        if value[11] is not None:
            # 遍历分割JSON提取_引用名称
            varStr = value[11]
            # 用分号分割varStr字符串，并保存到列表中
            varStrList = varStr.split(';')

            length = len(varStrList)
            # 遍历分割JSON表达式
            jsonStr = value[12]
            jsonList = jsonStr.split(';')

            # 循环输出列表值
            for i in range(length):
                # json引用变量名获取
                key = varStrList[i]
                # json表达式获取
                jsonExp = jsonList[i]
                # 字典值的获取
                valueJson = ak.get_text(res.text,jsonExp)
                # 变量和值的存储
                all_val[key] = valueJson



        try:
            # 结果校验
            # 实际结果
            # 给他一个默认值，让异常的报错信息好检查一些
            result = None
            result = ak.get_text(res.text, assert_value)
            print("============检查信息===========")
            print("校验字段： " + value[7])
            print("预期结果： " + value[8])
            print("实际结果： " + result)
            if result == expect_value:
                sheet.cell(r,10).value = '通过'
            else:
                sheet.cell(r, 10).value = '不通过'
            excel.save(excel_path)
            assert result == value[8]

        except:
            print("============实际结果=========")
            print("请求参数有误，请检查")
            sheet.cell(r, 10).value = '请求参数有误，请检查'
            excel.save(excel_path)

