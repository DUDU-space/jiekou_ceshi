#!/usr/bin/env python
# -*- coding: utf-8 -*-

# 打开excel
import openpyxl

# 读取excel
from autointf.kw.api_key import Apikey

excel = openpyxl.load_workbook('../data/api_cases.xlsx')
sheet = excel['Sheet1']

# 行数
r = 0
# 实例化工具类
ak = Apikey()

# 读取excel内容，实现文件驱动自动化执行
# 逐行循环读取Excel数据，value代表的是行
for value in sheet.values:
    r = r + 1
    print('行数： ' + str(r))
    # print(value)
    # print(type(value))
    # 判断当前行的第一列的值，是否是数字编号
    if type(value[0]) is int:
        # 准备需要的测试数据
        # 请求参数
        data = value[5]
        # print(data)
        # 校验字段
        assert_value = value[7]
        # print(assert_value)
        # 预期结果字段
        expect_value = value[8]
        # print(expect_value)
        # 如果存在请求头
        if value[4]:
            # print("存在请求头")
            # 存在请求参数
            if value[5]:
                # print("存在请求参数")
                dict_data = {
                    'url': value[1] + value[2],
                    # eval 将字符串str当做有效的表达式来求值并返回计算结果
                    # 这里直接给headers一个字典值
                    'headers':eval(value[4]),
                    value[6]: eval(data)
                }
                print(dict_data)
            # 不存在请求参数
            else:
                # print('不存在请求参数')
                dict_data = {
                    'url': value[1] + value[2],
                    'headers': eval(value[4])
                }
                print(dict_data)

        # 不存在请求头
        else:
            # print("不存在请求头")
            # 存在请求参数
            if value[5]:
                # print('存在请求参数')
                dict_data = {
                    'url': value[1] + value[2],
                    value[6]: eval(data)
                }
                print(dict_data)
            # 不存在请求参数
            else:
                # print('不存在请求参数')
                dict_data = {
                    'url': value[1] + value[2]
                }
                print(dict_data)

        """
            常规的参数传递：
            requests.post(url="",data="",headers="")
            封装时，参数可以直接用**kwargs
            requests.post(**kwargs)
        """
        # 使用反射模拟请求
        res = getattr(ak,value[3])(**dict_data)

        try:
            # 结果校验
            result = ak.get_text(res.text, assert_value)
            print("============实际结果=========")
            print(result == expect_value)
            if result == expect_value:
                sheet.cell(r,10).value = '通过'
            else:
                sheet.cell(r, 10).value = '不通过'
            excel.save('../data/api_cases.xlsx')

        except:
            print("============实际结果=========")
            print("请求参数有误，请检查")
            sheet.cell(r, 10).value = '请求参数有误，请检查'
            excel.save('../data/api_cases.xlsx')
